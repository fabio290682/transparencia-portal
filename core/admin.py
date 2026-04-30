from __future__ import annotations

import logging
from types import MethodType

logger = logging.getLogger(__name__)

_XLSX_CONTENT_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/octet-stream",
}
_MAX_IMPORT_FILE_SIZE = 50 * 1024 * 1024  # 50 MB

from django import forms
from django.contrib import admin, messages
from django.core.files.uploadedfile import UploadedFile
from django.db import OperationalError, ProgrammingError, transaction
from django.db.models import Count, Q
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect, render
from django.urls import NoReverseMatch, path, reverse

from .models import (
    Despesa,
    EsicPedido,
    Licitacao,
    PortalInformacao,
    Projeto,
    Servidor,
    UnidadeGestora,
)

admin.site.site_header = "Instituto Meio do Mundo"
admin.site.site_title = "Admin IMM"
admin.site.index_title = "Painel Administrativo"

_ESIC_STATUS_ABERTO = "ABERTO"

SECAO_MAP = {
    "FINANCEIROS": "FINANCEIROS",
    "FINANCEIRO": "FINANCEIROS",
    "PRESTACAO": "PRESTACAO",
    "PRESTAÇÃO": "PRESTACAO",
    "CONTRATACOES": "CONTRATACOES",
    "CONTRATAÇÕES": "CONTRATACOES",
    "POLITICAS": "POLITICAS",
    "POLÍTICAS": "POLITICAS",
}


class PortalInformacaoImportForm(forms.Form):
    arquivo = forms.FileField(
        label="Planilha Excel (.xlsx)",
        help_text="Colunas: secao, titulo, descricao, link, ordem, ativo",
    )


@admin.register(UnidadeGestora)
class UnidadeGestoraAdmin(admin.ModelAdmin):
    list_display = ("codigo", "nome", "sigla")
    search_fields = ("codigo", "nome", "sigla")
    readonly_fields = ("id",)


@admin.register(Despesa)
class DespesaAdmin(admin.ModelAdmin):
    list_display = ("codigo", "descricao", "categoria", "exercicio", "unidade")
    list_filter = ("categoria", "exercicio", "unidade")
    search_fields = ("codigo", "descricao")
    readonly_fields = ("id",)


@admin.register(Licitacao)
class LicitacaoAdmin(admin.ModelAdmin):
    list_display = ("numero", "modalidade", "status", "data_abertura", "unidade")
    list_filter = ("modalidade", "status", "unidade")
    search_fields = ("numero", "objeto")
    readonly_fields = ("id",)


@admin.register(Servidor)
class ServidorAdmin(admin.ModelAdmin):
    list_display = ("matricula", "nome", "cargo", "vinculo", "competencia", "unidade")
    list_filter = ("vinculo", "competencia", "unidade")
    search_fields = ("matricula", "nome", "cargo")
    readonly_fields = ("id",)


@admin.register(EsicPedido)
class EsicPedidoAdmin(admin.ModelAdmin):
    list_display = ("protocolo", "tipo", "status", "email", "prazo", "unidade")
    list_filter = ("tipo", "status", "unidade")
    search_fields = ("protocolo", "descricao", "email")
    readonly_fields = ("id",)


@admin.register(PortalInformacao)
class PortalInformacaoAdmin(admin.ModelAdmin):
    change_list_template = "admin/core/portalinformacao/change_list.html"
    list_display = (
        "secao",
        "titulo",
        "tipo_documento",
        "ordem",
        "ativo",
        "atualizado_em",
    )
    list_filter = ("secao", "ativo")
    search_fields = ("titulo", "descricao", "link")
    list_editable = ("ordem", "ativo")
    list_display_links = ("titulo",)
    readonly_fields = ("id", "tipo_documento", "criado_em", "atualizado_em")
    ordering = ("secao", "ordem", "titulo")

    fieldsets = (
        ("Classificação", {"fields": ("secao", "titulo", "descricao", "ordem", "ativo")}),
        (
            "Documento para o Portal",
            {
                "fields": ("link", "tipo_documento"),
                "description": (
                    "Cole aqui o link do documento. "
                    "No Google Drive: abra o arquivo → Compartilhar → "
                    "'Qualquer pessoa com o link' → copiar link."
                ),
            },
        ),
        ("Controle", {"fields": ("id", "criado_em", "atualizado_em")}),
    )

    def get_urls(self) -> list:
        urls = super().get_urls()
        custom_urls = [
            path(
                "importar-planilha/",
                self.admin_site.admin_view(self.importar_planilha_view),
                name="core_portalinformacao_importar_planilha",
            ),
        ]
        return custom_urls + urls

    def _to_bool(self, value: object) -> bool:
        if isinstance(value, bool):
            return value
        if value is None:
            return True
        return str(value).strip().lower() in {"1", "true", "sim", "s", "yes", "y", "ativo"}

    def _normalizar_secao(self, secao: object) -> str | None:
        if secao is None:
            return None
        return SECAO_MAP.get(str(secao).strip().upper())

    @transaction.atomic
    def _importar_planilha(self, arquivo: UploadedFile) -> int:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise RuntimeError(
                "Dependencia ausente: instale openpyxl para usar a importacao de planilhas."
            ) from exc

        wb = load_workbook(filename=arquivo, data_only=True, read_only=True)
        ws = wb.active

        header = [
            str(c).strip().lower() if c is not None else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        ]
        header_map = {name: idx for idx, name in enumerate(header)}

        required = {"secao", "titulo", "descricao"}
        missing = sorted(required - set(header_map.keys()))
        if missing:
            raise ValueError(f'Colunas obrigatorias ausentes: {", ".join(missing)}')

        objects: list[PortalInformacao] = []
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            secao_raw = row[header_map["secao"]] if "secao" in header_map else None
            secao = self._normalizar_secao(secao_raw)
            if not secao:
                raise ValueError(f"Seção inválida na linha {row_index}: {secao_raw}")

            titulo_raw = row[header_map["titulo"]]
            titulo = str(titulo_raw).strip() if titulo_raw is not None else ""
            descricao_raw = row[header_map["descricao"]]
            descricao = str(descricao_raw).strip() if descricao_raw is not None else ""
            if not titulo or not descricao:
                raise ValueError(f"Titulo e descricao sao obrigatorios na linha {row_index}")

            link = ""
            if "link" in header_map:
                raw_link = row[header_map["link"]]
                link = str(raw_link).strip() if raw_link is not None else ""

            ordem = 0
            if "ordem" in header_map:
                raw_ordem = row[header_map["ordem"]]
                if raw_ordem not in (None, ""):
                    try:
                        ordem = int(raw_ordem)
                    except (TypeError, ValueError) as exc:
                        raise ValueError(
                            f"Ordem invalida na linha {row_index}: {raw_ordem}"
                        ) from exc

            ativo = True
            if "ativo" in header_map:
                ativo = self._to_bool(row[header_map["ativo"]])

            objects.append(
                PortalInformacao(
                    secao=secao,
                    titulo=titulo,
                    descricao=descricao,
                    link=link,
                    ordem=ordem,
                    ativo=ativo,
                )
            )

        PortalInformacao.objects.bulk_create(objects, batch_size=500)
        return len(objects)

    def importar_planilha_view(self, request: HttpRequest) -> HttpResponse:
        if request.method == "POST":
            form = PortalInformacaoImportForm(request.POST, request.FILES)
            if form.is_valid():
                arquivo = form.cleaned_data["arquivo"]
                if arquivo.size > _MAX_IMPORT_FILE_SIZE:
                    form.add_error("arquivo", "Arquivo muito grande. Máximo permitido: 50 MB.")
                elif not arquivo.name.lower().endswith(".xlsx"):
                    form.add_error("arquivo", "Formato invalido. Envie um arquivo .xlsx")
                elif arquivo.content_type not in _XLSX_CONTENT_TYPES:
                    form.add_error("arquivo", "Tipo de arquivo inválido. Envie um .xlsx real.")
                else:
                    try:
                        created = self._importar_planilha(arquivo)
                    except ValueError as exc:
                        form.add_error(None, f"Dados inválidos na planilha: {exc}")
                    except RuntimeError as exc:
                        logger.error("Falha ao importar planilha: %s", exc, exc_info=True)
                        form.add_error(None, "Erro ao processar o arquivo. Verifique o formato e tente novamente.")
                    else:
                        self.message_user(
                            request,
                            f"Importacao concluida com sucesso. Registros criados: {created}",
                            level=messages.SUCCESS,
                        )
                        return redirect("admin:core_portalinformacao_changelist")
        else:
            form = PortalInformacaoImportForm()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.opts,
            "title": "Importar Planilha (Portal Informacoes)",
            "form": form,
        }
        return render(request, "admin/core/portalinformacao/importar_planilha.html", context)

    @admin.display(description="Tipo de Documento")
    def tipo_documento(self, obj: PortalInformacao) -> str:
        return "Link Externo" if obj.link else "Sem documento"


@admin.register(Projeto)
class ProjetoAdmin(admin.ModelAdmin):
    list_display = (
        "status",
        "titulo",
        "tipo_documento",
        "valor",
        "ativo",
        "ordem",
        "atualizado_em",
    )
    list_filter = ("status", "ativo")
    search_fields = ("titulo", "subtitulo", "instrumento", "nome_projeto", "autoria_emenda")
    list_editable = ("ordem", "ativo")
    list_display_links = ("titulo",)
    readonly_fields = ("id", "tipo_documento", "criado_em", "atualizado_em")
    ordering = ("status", "ordem", "titulo")
    fieldsets = (
        ("Publicacao", {"fields": ("status", "titulo", "subtitulo", "icone", "ordem", "ativo")}),
        (
            "Financeiro e Instrumento",
            {
                "fields": (
                    "valor",
                    "instrumento",
                    "recurso_financeiro",
                    "fonte_recurso_estadual",
                    "fonte_recurso_federal",
                    "autoria_emenda",
                    "valor_repassado",
                    "valor_global",
                ),
            },
        ),
        (
            "Detalhes do Projeto",
            {
                "fields": (
                    "nome_projeto",
                    "objeto",
                    "data_inicio",
                    "data_finalizacao",
                    "justificativa",
                    "dados_orcamentarios",
                ),
            },
        ),
        (
            "Documento para o Portal",
            {
                "fields": ("documento_link", "documento_label", "tipo_documento"),
                "description": (
                    "Cole aqui o link do documento (PDF, planilha, etc). "
                    "No Google Drive: abra o arquivo → Compartilhar → "
                    "'Qualquer pessoa com o link' → copiar link. "
                    "Em 'Rótulo do documento' escreva o nome que vai aparecer no botão (ex: 'Ver Relatório')."
                ),
            },
        ),
        ("Controle", {"fields": ("id", "criado_em", "atualizado_em")}),
    )

    @admin.display(description="Tipo de Documento")
    def tipo_documento(self, obj: Projeto) -> str:
        return "Link Externo" if obj.documento_link else "Sem documento"


def _safe_reverse(viewname: str) -> str:
    try:
        return reverse(viewname)
    except NoReverseMatch:
        return "#"


_dashboard_links_cache: list | None = None


def _get_dashboard_links() -> list:
    global _dashboard_links_cache
    if _dashboard_links_cache is None:
        _dashboard_links_cache = [
            {
                "label": "Projetos",
                "hint": "Cadastre, edite e organize os termos e projetos do portal.",
                "url": _safe_reverse("admin:core_projeto_changelist"),
                "add_url": _safe_reverse("admin:core_projeto_add"),
            },
            {
                "label": "Documentos",
                "hint": "Gerencie arquivos e links das áreas de transparência.",
                "url": _safe_reverse("admin:core_portalinformacao_changelist"),
                "add_url": _safe_reverse("admin:core_portalinformacao_add"),
            },
            {
                "label": "e-SIC",
                "hint": "Acompanhe protocolos, respostas e status dos pedidos.",
                "url": _safe_reverse("admin:core_esicpedido_changelist"),
                "add_url": _safe_reverse("admin:core_esicpedido_add"),
            },
            {
                "label": "Importar Planilha",
                "hint": "Faça carga em lote de documentos do portal.",
                "url": _safe_reverse("admin:core_portalinformacao_importar_planilha"),
                "add_url": _safe_reverse("admin:core_portalinformacao_importar_planilha"),
            },
            {
                "label": "Usuários",
                "hint": "Controle quem pode acessar e administrar o sistema.",
                "url": _safe_reverse("admin:auth_user_changelist"),
                "add_url": _safe_reverse("admin:auth_user_add"),
            },
            {
                "label": "Portal Público",
                "hint": "Abra o portal publicado para revisar as alterações.",
                "url": "/",
                "add_url": "/",
            },
        ]
    return _dashboard_links_cache


def _build_dashboard_context() -> dict:
    try:
        recent_projects = list(Projeto.objects.order_by("-atualizado_em")[:5])
        recent_documents = list(PortalInformacao.objects.order_by("-atualizado_em")[:5])
        recent_esic = list(EsicPedido.objects.order_by("-prazo")[:5])

        esic_counts = EsicPedido.objects.aggregate(
            total=Count("id"),
            pending=Count("id", filter=Q(status=_ESIC_STATUS_ABERTO)),
        )
        project_count = Projeto.objects.count()
        document_count = PortalInformacao.objects.count()
        esic_count = esic_counts["total"] or 0
        pending_esic_count = esic_counts["pending"] or 0
    except (OperationalError, ProgrammingError):
        recent_projects = []
        recent_documents = []
        recent_esic = []
        project_count = 0
        document_count = 0
        esic_count = 0
        pending_esic_count = 0

    return {
        "project_count": project_count,
        "document_count": document_count,
        "esic_count": esic_count,
        "pending_esic_count": pending_esic_count,
        "recent_projects": recent_projects,
        "recent_documents": recent_documents,
        "recent_esic": recent_esic,
        "portal_public_url": "/",
        "dashboard_links": _get_dashboard_links(),
    }


_original_admin_index = admin.site.index


def _custom_admin_index(
    self: admin.AdminSite,
    request: HttpRequest,
    extra_context: dict | None = None,
) -> HttpResponse:
    context = _build_dashboard_context()
    if extra_context:
        context.update(extra_context)
    return _original_admin_index(request, extra_context=context)


admin.site.index = MethodType(_custom_admin_index, admin.site)
